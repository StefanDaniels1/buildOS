"""
SDK Tools: MCP Tool Registration for buildOS

Registers IFC analysis tools as MCP server for SDK access.
Tools are namespaced as: mcp__ifc__<tool_name>
"""

from claude_agent_sdk import tool, create_sdk_mcp_server
from typing import Dict, Any
import json
import os
from pathlib import Path


def _resolve_path(path_str: str) -> Path:
    """
    Resolve paths, handling environment variables and relative paths.

    Handles:
    - $CLAUDE_PROJECT_DIR -> actual project directory
    - /app/ prefix -> project directory (Docker compatibility)
    - Relative paths -> absolute paths
    """
    if not path_str:
        return Path(path_str)

    # Replace $CLAUDE_PROJECT_DIR with actual project directory
    if "$CLAUDE_PROJECT_DIR" in path_str:
        project_dir = str(Path(__file__).parent.resolve())
        path_str = path_str.replace("$CLAUDE_PROJECT_DIR", project_dir)

    # Handle /app/ prefix (Docker compatibility)
    if path_str.startswith("/app/"):
        project_dir = Path(__file__).parent.resolve()
        path_str = str(project_dir / path_str[5:])

    # Expand user and environment variables
    path_str = os.path.expanduser(os.path.expandvars(path_str))

    return Path(path_str)


@tool(
    name="parse_ifc_file",
    description="Parse IFC file and extract building element data to JSON format. Paths are auto-resolved.",
    input_schema={
        "ifc_path": str,
        "output_path": str
    }
)
async def parse_ifc_file(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Parse IFC file using ifcopenshell.

    Args:
        ifc_path: Path to input IFC file
        output_path: Path to output JSON file

    Returns:
        Success status and entity count
    """
    try:
        # Import IFC parser from tools
        import sys
        tools_path = Path(__file__).parent / ".claude" / "tools"
        sys.path.insert(0, str(tools_path))

        # Resolve paths (handles $CLAUDE_PROJECT_DIR, /app/, etc.)
        ifc_path = _resolve_path(args["ifc_path"])
        output_path_resolved = _resolve_path(args["output_path"])

        # Validate IFC file exists
        if not ifc_path.exists():
            return {
                "content": [{
                    "type": "text",
                    "text": json.dumps({
                        "success": False,
                        "error": f"IFC file not found: {ifc_path}",
                        "original_path": args["ifc_path"]
                    }, indent=2)
                }],
                "is_error": True
            }

        # Use existing IFC parser from agent_system4
        import ifcopenshell

        ifc_file = ifcopenshell.open(str(ifc_path))

        # Extract all rooted entities (elements that have geometry)
        elements = []
        for element in ifc_file.by_type("IfcProduct"):
            if element.is_a() in [
                "IfcBeam", "IfcColumn", "IfcWall", "IfcSlab", "IfcDoor", "IfcWindow",
                "IfcRoof", "IfcStair", "IfcCovering", "IfcFurnishingElement", "IfcFooting"
            ]:
                elements.append({
                    "guid": element.GlobalId,
                    "ifc_type": element.is_a(),
                    "name": element.Name if hasattr(element, 'Name') else None,
                    "object_type": element.ObjectType if hasattr(element, 'ObjectType') else None,
                    "description": element.Description if hasattr(element, 'Description') else None
                })

        # Write to JSON
        output_path_resolved.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path_resolved, 'w') as f:
            json.dump({
                "elements": elements,
                "total_count": len(elements),
                "source_file": str(ifc_path)
            }, f, indent=2)

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": True,
                    "entities_parsed": len(elements),
                    "output_file": str(output_path_resolved)
                }, indent=2)
            }]
        }

    except Exception as e:
        import traceback
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="prepare_batches",
    description="Create element batches for parallel classification processing",
    input_schema={
        "json_path": str,
        "batch_size": int,
        "output_path": str
    }
)
async def prepare_batches(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Split elements into batches for parallel processing.

    Args:
        json_path: Path to parsed IFC JSON
        batch_size: Elements per batch (default: 100)
        output_path: Path to output batches.json

    Returns:
        Success status and batch count
    """
    try:
        # Load parsed elements
        with open(args["json_path"], 'r') as f:
            data = json.load(f)

        elements = data.get("elements", [])
        batch_size = args.get("batch_size", 100)

        # Create batches
        batches = []
        for i in range(0, len(elements), batch_size):
            batch = {
                "batch_id": len(batches) + 1,
                "elements": elements[i:i + batch_size],
                "element_count": len(elements[i:i + batch_size])
            }
            batches.append(batch)

        # Write batches
        output_path = Path(args["output_path"])
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, 'w') as f:
            json.dump({
                "batches": batches,
                "total_batches": len(batches),
                "total_elements": len(elements),
                "batch_size": batch_size
            }, f, indent=2)

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": True,
                    "total_batches": len(batches),
                    "total_elements": len(elements),
                    "output_file": str(output_path)
                }, indent=2)
            }]
        }

    except Exception as e:
        import traceback
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="calculate_co2",
    description="Calculate CO2 impact from classified building elements using durability database. Paths are auto-resolved (supports $CLAUDE_PROJECT_DIR and /app/ prefixes).",
    input_schema={
        "classified_path": str,
        "database_path": str,
        "output_path": str
    }
)
async def calculate_co2(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Calculate CO2 emissions for classified elements.

    Args:
        classified_path: Path to classified elements JSON
        database_path: Path to durability database JSON (auto-resolved)
        output_path: Path to output CO2 results

    Returns:
        Success status and total CO2
    """
    try:
        # Resolve paths (handles $CLAUDE_PROJECT_DIR, /app/, etc.)
        classified_path = _resolve_path(args["classified_path"])
        database_path = _resolve_path(args["database_path"])
        output_path = _resolve_path(args["output_path"])

        # Validate input files exist
        if not classified_path.exists():
            return {
                "content": [{
                    "type": "text",
                    "text": json.dumps({
                        "success": False,
                        "error": f"Classified elements file not found: {classified_path}",
                        "hint": "Make sure aggregate_batch_results was called first"
                    }, indent=2)
                }],
                "is_error": True
            }

        if not database_path.exists():
            # Try default database location
            default_db = Path(__file__).parent / ".claude" / "skills" / "ifc-analysis" / "reference" / "durability_database.json"
            if default_db.exists():
                database_path = default_db
            else:
                return {
                    "content": [{
                        "type": "text",
                        "text": json.dumps({
                            "success": False,
                            "error": f"Durability database not found: {database_path}",
                            "hint": f"Default location: {default_db}"
                        }, indent=2)
                    }],
                    "is_error": True
                }

        # Load classified elements
        with open(classified_path, 'r') as f:
            classified_data = json.load(f)

        # Load durability database
        with open(database_path, 'r') as f:
            database = json.load(f)

        # Get materials lookup from database
        materials_db = database.get("materials", database)

        # Calculate CO2 for each element
        results = []
        total_co2 = 0.0
        elements_with_volume = 0
        elements_without_volume = 0
        by_category = {}

        for element in classified_data.get("elements", []):
            # Get material classification
            material_primary = element.get("material_primary", {})
            if isinstance(material_primary, dict):
                category = material_primary.get("category", "unknown")
                subcategory = material_primary.get("subcategory", "generic")
            else:
                category = str(material_primary) if material_primary else "unknown"
                subcategory = "generic"

            # Get volume (handle null/None gracefully)
            volume = element.get("volume_m3")
            if volume is None or volume == "null":
                volume = 0.0
                elements_without_volume += 1
            else:
                try:
                    volume = float(volume)
                    elements_with_volume += 1
                except (ValueError, TypeError):
                    volume = 0.0
                    elements_without_volume += 1

            # Look up CO2 factor in database
            co2_factor = 0.0
            material_info = materials_db.get(category, {})
            if isinstance(material_info, dict):
                sub_info = material_info.get(subcategory, material_info.get(f"{category}_generic", {}))
                if isinstance(sub_info, dict):
                    # Calculate CO2: embodied_co2_per_kg * density * volume
                    embodied_co2 = sub_info.get("embodied_co2_per_kg", 0.0)
                    density = sub_info.get("density_kg_m3", 2400)  # default concrete density
                    co2_factor = embodied_co2 * density

            element_co2 = volume * co2_factor

            # Track by category
            if category not in by_category:
                by_category[category] = {"count": 0, "co2_kg": 0.0, "volume_m3": 0.0}
            by_category[category]["count"] += 1
            by_category[category]["co2_kg"] += element_co2
            by_category[category]["volume_m3"] += volume

            results.append({
                **element,
                "co2_kg": round(element_co2, 2),
                "co2_factor_per_m3": round(co2_factor, 2)
            })

            total_co2 += element_co2

        # Calculate percentages
        for cat in by_category:
            by_category[cat]["percentage"] = round(by_category[cat]["co2_kg"] / total_co2 * 100, 1) if total_co2 > 0 else 0

        # Write results
        output_path.parent.mkdir(parents=True, exist_ok=True)

        report = {
            "summary": {
                "total_co2_kg": round(total_co2, 2),
                "element_count": len(results),
                "elements_with_volume": elements_with_volume,
                "elements_without_volume": elements_without_volume,
                "completeness_pct": round(elements_with_volume / len(results) * 100, 1) if results else 0
            },
            "by_category": by_category,
            "elements": results
        }

        with open(output_path, 'w') as f:
            json.dump(report, f, indent=2)

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": True,
                    "total_co2_kg": round(total_co2, 2),
                    "element_count": len(results),
                    "elements_with_volume": elements_with_volume,
                    "elements_without_volume": elements_without_volume,
                    "completeness_pct": report["summary"]["completeness_pct"],
                    "output_file": str(output_path)
                }, indent=2)
            }]
        }

    except Exception as e:
        import traceback
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="generate_excel_report",
    description="WARNING: Do NOT use this tool unless user explicitly says 'download Excel', 'export to xlsx', or 'Excel file'. For displaying tabular data, use generate_spreadsheet instead. This tool creates a downloadable .xlsx file - only use when user specifically requests a file download.",
    input_schema={
        "prompt": str,
        "output_dir": str,
        "data_json": str  # Optional: JSON string OR file path to JSON file
    }
)
async def generate_excel_report(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generate an Excel report using Claude's Excel skill.

    Args:
        prompt: Description of what Excel file to create
        output_dir: Directory to save the output file
        data_json: Optional JSON string OR file path to a JSON file

    Returns:
        Success status and file path
    """
    try:
        import os

        # First, try to use the Skills API if available
        api_key = os.environ.get("ANTHROPIC_API_KEY")

        # Resolve output directory path
        output_dir = _resolve_path(args["output_dir"])

        # Load data from file path or parse JSON string
        data_str = ""
        if args.get("data_json"):
            data_json_input = args["data_json"]

            # Check if it's a file path (resolve it first)
            if data_json_input.startswith("/") or data_json_input.startswith("./") or data_json_input.startswith("$") or data_json_input.startswith("workspace"):
                resolved_path = _resolve_path(data_json_input)
                try:
                    with open(resolved_path, 'r') as f:
                        data = json.load(f)
                        data_str = json.dumps(data, indent=2)
                except FileNotFoundError:
                    return {
                        "content": [{
                            "type": "text",
                            "text": json.dumps({
                                "success": False,
                                "error": f"Data file not found: {resolved_path}",
                                "original_path": data_json_input
                            }, indent=2)
                        }],
                        "is_error": True
                    }
                except json.JSONDecodeError as e:
                    return {
                        "content": [{
                            "type": "text",
                            "text": json.dumps({
                                "success": False,
                                "error": f"Invalid JSON in file {resolved_path}: {str(e)}"
                            }, indent=2)
                        }],
                        "is_error": True
                    }
            else:
                # Try to parse as JSON string
                try:
                    data = json.loads(data_json_input)
                    data_str = json.dumps(data, indent=2)
                except json.JSONDecodeError:
                    # Use as-is if not valid JSON
                    data_str = data_json_input
        
        # Try Skills API first (if anthropic package is available and API key is set)
        if api_key:
            try:
                from skills_client import SkillsClient
                
                client = SkillsClient(api_key=api_key)
                
                # Build prompt with data
                full_prompt = args["prompt"]
                if data_str:
                    full_prompt = f"""Using the following data:

```json
{data_str}
```

{args["prompt"]}
"""
                
                result = await client.generate_excel(
                    prompt=full_prompt,
                    output_dir=str(output_dir)
                )
                
                if result["success"] and result["files"]:
                    # Generate download URLs for Skills API files
                    download_urls = []
                    for f in result["files"]:
                        f_str = str(f)
                        if f_str.startswith("/app/"):
                            download_path = f_str[5:]
                        else:
                            download_path = f_str
                        url = f"/workspace/{download_path.replace('workspace/', '')}" if download_path.startswith("workspace/") else f"/{download_path}"
                        download_urls.append(url)

                    return {
                        "content": [{
                            "type": "text",
                            "text": json.dumps({
                                "success": True,
                                "files": result["files"],
                                "download_urls": download_urls,
                                "response": result["response"][:500] if result["response"] else "",
                                "method": "claude_skills_api"
                            }, indent=2)
                        }]
                    }
                else:
                    # Skills API failed, fall through to openpyxl fallback
                    print(f"Skills API failed: {result.get('error')}, trying openpyxl fallback")
                    
            except ImportError:
                print("anthropic package not available, using openpyxl fallback")
            except Exception as e:
                print(f"Skills API error: {e}, using openpyxl fallback")
        
        # Fallback: Use openpyxl to create Excel file locally
        # Pass resolved output_dir in args
        args_with_resolved = {**args, "output_dir": str(output_dir)}
        return await _generate_excel_with_openpyxl(args_with_resolved, data_str)
        
    except Exception as e:
        import traceback
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }, indent=2)
            }],
            "is_error": True
        }


async def _generate_excel_with_openpyxl(args: Dict[str, Any], data_str: str) -> Dict[str, Any]:
    """
    Fallback Excel generation using openpyxl.
    Creates a basic Excel file with the provided data.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from pathlib import Path
        from datetime import datetime
        
        # Parse data if available
        data = None
        if data_str:
            try:
                data = json.loads(data_str)
            except json.JSONDecodeError:
                pass
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        
        # Add title
        ws['A1'] = "Generated Report"
        ws['A1'].font = title_font
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Prompt: {args['prompt'][:100]}..."
        
        row = 5
        
        # If we have structured data, try to render it
        if data and isinstance(data, dict):
            # Handle CO2 report format
            if 'summary' in data:
                ws[f'A{row}'] = "Summary"
                ws[f'A{row}'].font = title_font
                row += 1
                
                summary = data['summary']
                for key, value in summary.items():
                    ws[f'A{row}'] = key.replace('_', ' ').title()
                    ws[f'B{row}'] = str(value)
                    row += 1
                row += 1
            
            # Handle by_category
            if 'by_category' in data:
                ws[f'A{row}'] = "By Category"
                ws[f'A{row}'].font = title_font
                row += 1

                # Determine available columns from first category
                first_cat = next(iter(data['by_category'].values()), {})
                has_mass = 'mass_kg' in first_cat
                has_volume = 'volume_m3' in first_cat

                headers = ["Category", "Count", "CO2 (kg)"]
                if has_mass:
                    headers.append("Mass (kg)")
                if has_volume:
                    headers.append("Volume (m³)")
                headers.append("Percentage")

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                row += 1

                for category, details in data['by_category'].items():
                    col = 1
                    ws.cell(row=row, column=col).value = category.replace('_', ' ').title()
                    col += 1
                    ws.cell(row=row, column=col).value = details.get('count', 0)
                    col += 1
                    ws.cell(row=row, column=col).value = details.get('co2_kg', 0)
                    col += 1
                    if has_mass:
                        ws.cell(row=row, column=col).value = details.get('mass_kg', 0)
                        col += 1
                    if has_volume:
                        ws.cell(row=row, column=col).value = details.get('volume_m3', 0)
                        col += 1
                    ws.cell(row=row, column=col).value = details.get('percentage', 0)
                    row += 1
                row += 1
            
            # Handle detailed_results OR elements (CO2 reports use 'elements')
            results_key = 'detailed_results' if 'detailed_results' in data else 'elements'
            if results_key in data and data[results_key]:
                ws[f'A{row}'] = "Detailed Results" if results_key == 'detailed_results' else "Elements"
                ws[f'A{row}'].font = title_font
                row += 1

                # Get headers from first item, prioritize important columns
                first_item = data[results_key][0]
                all_keys = list(first_item.keys())

                # Prioritize these columns for BIM data
                priority_keys = ['name', 'ifc_type', 'element_type', 'material_primary',
                                'volume_m3', 'co2_kg', 'mass_kg', 'confidence', 'global_id']
                headers = [k for k in priority_keys if k in all_keys]
                headers.extend([k for k in all_keys if k not in headers and k not in ['reasoning']])
                headers = headers[:12]  # Limit to 12 columns

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = header.replace('_', ' ').title()
                    cell.font = header_font
                    cell.fill = header_fill
                row += 1

                for item in data[results_key]:
                    for col, key in enumerate(headers, 1):
                        value = item.get(key, '')
                        # Handle nested dicts (like material_primary)
                        if isinstance(value, dict):
                            value = value.get('category', str(value))
                        ws.cell(row=row, column=col).value = str(value) if value is not None else ''
                    row += 1
        
        # Adjust column widths
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Save file
        output_dir = Path(args["output_dir"])
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = output_dir / f"report_{timestamp}.xlsx"
        wb.save(str(output_file))

        # Generate download URL (strip /app prefix, use /workspace route)
        output_str = str(output_file)
        if output_str.startswith("/app/"):
            download_path = output_str[5:]  # Remove "/app/" prefix
        else:
            download_path = output_str
        download_url = f"/workspace/{download_path.replace('workspace/', '')}" if download_path.startswith("workspace/") else f"/{download_path}"

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": True,
                    "files": [str(output_file)],
                    "download_url": download_url,
                    "response": f"Excel file created. Download: {download_url}",
                    "method": "openpyxl_fallback"
                }, indent=2)
            }]
        }
        
    except ImportError:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": "Neither anthropic (for Skills API) nor openpyxl (for local generation) is available. Install one of: pip install anthropic OR pip install openpyxl"
                }, indent=2)
            }],
            "is_error": True
        }
    except Exception as e:
        import traceback
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="generate_spreadsheet",
    description="PREFERRED tool for all tabular data output (bill of materials, element lists, CO2 breakdowns, etc). Displays data in the interactive Spreadsheet Builder where users can review, edit, and export. Use this INSTEAD of generate_excel_report for initial data display. Only use generate_excel_report when user explicitly requests a downloadable Excel file. The data parameter accepts a 2D array like [[\"Header1\", \"Header2\"], [\"row1col1\", \"row1col2\"]].",
    input_schema={
        "name": str,
        "data": str,  # 2D array as JSON string OR actual array - will be parsed
        "columns": str  # Column definitions as JSON string OR actual array - optional
    }
)
async def generate_spreadsheet(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generate spreadsheet data for the interactive frontend editor.

    Args:
        name: Name/title of the spreadsheet
        data: 2D array of cell values (rows x columns) - can be array or JSON string
        columns: List of column definitions (e.g., [{"title": "Element", "width": 150}, ...])

    Returns:
        Spreadsheet data in jspreadsheet-compatible format
    """
    try:
        name = args.get("name", "Spreadsheet")
        data = args.get("data", [])
        columns = args.get("columns", [])

        # Handle data passed as JSON string (common when agent constructs it)
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except json.JSONDecodeError:
                return {
                    "content": [{
                        "type": "text",
                        "text": json.dumps({
                            "success": False,
                            "error": "Invalid data format. Expected a 2D array or valid JSON string.",
                            "received": data[:200] if len(data) > 200 else data
                        }, indent=2)
                    }],
                    "is_error": True
                }

        # Handle columns passed as JSON string
        if isinstance(columns, str):
            try:
                columns = json.loads(columns)
            except json.JSONDecodeError:
                columns = []

        # Validate data is a 2D array
        if not isinstance(data, list):
            return {
                "content": [{
                    "type": "text",
                    "text": json.dumps({
                        "success": False,
                        "error": "Data must be a 2D array (list of lists)"
                    }, indent=2)
                }],
                "is_error": True
            }

        # Ensure columns are properly formatted
        formatted_columns = []
        for col in columns:
            if isinstance(col, str):
                formatted_columns.append({"title": col, "width": 120})
            elif isinstance(col, dict):
                formatted_columns.append({
                    "title": col.get("title", ""),
                    "width": col.get("width", 120),
                    "type": col.get("type", "text")
                })
            else:
                formatted_columns.append({"title": str(col), "width": 120})

        # If no columns provided, generate from first row or default
        if not formatted_columns and data and len(data) > 0:
            first_row = data[0] if isinstance(data[0], list) else [data[0]]
            formatted_columns = [{"title": f"Column {i+1}", "width": 120} for i in range(len(first_row))]

        spreadsheet_data = {
            "type": "spreadsheet",
            "name": name,
            "data": data,
            "columns": formatted_columns,
            "row_count": len(data),
            "column_count": len(formatted_columns)
        }

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": True,
                    "spreadsheet": spreadsheet_data,
                    "message": f"Spreadsheet '{name}' generated with {len(data)} rows and {len(formatted_columns)} columns. Switch to the Spreadsheet view to see and edit the data."
                }, indent=2)
            }]
        }

    except Exception as e:
        import traceback
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="generate_presentation",
    description="Generate a PowerPoint presentation using Claude's PowerPoint skill. Use this for creating professional presentations with slides, charts, and formatting. The data_json parameter can be either a JSON string or a file path to a JSON file.",
    input_schema={
        "prompt": str,
        "output_dir": str,
        "data_json": str  # Optional: JSON string OR file path to JSON file
    }
)
async def generate_presentation(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generate a PowerPoint presentation using Claude's PPTX skill.
    
    Args:
        prompt: Description of what presentation to create
        output_dir: Directory to save the output file
        data_json: Optional JSON string OR file path to a JSON file
        
    Returns:
        Success status and file path
    """
    try:
        import os
        
        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            return {
                "content": [{
                    "type": "text",
                    "text": json.dumps({
                        "success": False,
                        "error": "ANTHROPIC_API_KEY not set. Required for PowerPoint skill. Note: PowerPoint generation requires the Claude Skills API (no local fallback available)."
                    }, indent=2)
                }],
                "is_error": True
            }
        
        # Load data from file path or parse JSON string
        data_str = ""
        if args.get("data_json"):
            data_json_input = args["data_json"]
            
            # Check if it's a file path
            if data_json_input.startswith("/") or data_json_input.startswith("./"):
                try:
                    with open(data_json_input, 'r') as f:
                        data = json.load(f)
                        data_str = json.dumps(data, indent=2)
                except FileNotFoundError:
                    return {
                        "content": [{
                            "type": "text",
                            "text": json.dumps({
                                "success": False,
                                "error": f"Data file not found: {data_json_input}"
                            }, indent=2)
                        }],
                        "is_error": True
                    }
                except json.JSONDecodeError as e:
                    return {
                        "content": [{
                            "type": "text",
                            "text": json.dumps({
                                "success": False,
                                "error": f"Invalid JSON in file {data_json_input}: {str(e)}"
                            }, indent=2)
                        }],
                        "is_error": True
                    }
            else:
                # Try to parse as JSON string
                try:
                    data = json.loads(data_json_input)
                    data_str = json.dumps(data, indent=2)
                except json.JSONDecodeError:
                    data_str = data_json_input
        
        from skills_client import SkillsClient
        
        client = SkillsClient(api_key=api_key)
        
        full_prompt = args["prompt"]
        if data_str:
            full_prompt = f"""Using the following data:

```json
{data_str}
```

{args["prompt"]}
"""
        
        result = await client.generate_presentation(
            prompt=full_prompt,
            output_dir=args["output_dir"]
        )
        
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": result["success"],
                    "files": result["files"],
                    "response": result["response"][:500] if result["response"] else "",
                    "error": result.get("error")
                }, indent=2)
            }],
            "is_error": not result["success"]
        }
        
    except ImportError:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": "anthropic package not installed. PowerPoint generation requires the Claude Skills API. Install with: pip install anthropic"
                }, indent=2)
            }],
            "is_error": True
        }
    except Exception as e:
        import traceback
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="generate_pdf_report",
    description="Generate a PDF sustainability report using ReportLab. This creates a professional PDF directly without external APIs. Use this for CO2 analysis reports. Requires a co2_report.json file as input. Paths are auto-resolved.",
    input_schema={
        "co2_report_path": str,
        "ifc_filename": str,
        "output_path": str
    }
)
async def generate_pdf_report(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Generate a PDF sustainability report using ReportLab (local generation, no API).

    Args:
        co2_report_path: Path to the CO2 report JSON file (e.g., co2_report.json)
        ifc_filename: Original IFC filename (for report title)
        output_path: Where to save the PDF file

    Returns:
        Success status and file path
    """
    try:
        import sys

        # Resolve paths (handles $CLAUDE_PROJECT_DIR, /app/, etc.)
        co2_report_path = _resolve_path(args["co2_report_path"])
        output_path = _resolve_path(args["output_path"])
        ifc_filename = args["ifc_filename"]

        # Validate input file exists
        if not co2_report_path.exists():
            return {
                "content": [{
                    "type": "text",
                    "text": json.dumps({
                        "success": False,
                        "error": f"CO2 report file not found: {co2_report_path}",
                        "hint": "Make sure calculate_co2 was called first and produced co2_report.json"
                    }, indent=2)
                }],
                "is_error": True
            }

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Import the ReportLab-based PDF generator
        tools_path = Path(__file__).parent / ".claude" / "tools"
        sys.path.insert(0, str(tools_path))

        try:
            from generate_co2_pdf import generate_co2_report_pdf
        except ImportError:
            # Fallback: try to import from skills path
            skills_path = Path(__file__).parent / ".claude" / "skills" / "ifc-analysis" / "scripts"
            sys.path.insert(0, str(skills_path))
            from generate_pdf import generate_co2_report_pdf

        # Generate PDF using ReportLab
        result = generate_co2_report_pdf(
            co2_report_file=str(co2_report_path),
            ifc_filename=ifc_filename,
            output_pdf=str(output_path)
        )

        # Generate download URL (strip /app prefix, use /workspace route)
        output_str = str(output_path)
        if output_str.startswith("/app/"):
            download_path = output_str[5:]  # Remove "/app/" prefix
        else:
            download_path = output_str
        download_url = f"/workspace/{download_path.replace('workspace/', '')}" if download_path.startswith("workspace/") else f"/{download_path}"

        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": True,
                    "pdf_file": str(output_path),
                    "download_url": download_url,
                    "total_co2_kg": result.get("total_co2_kg"),
                    "completeness_pct": result.get("completeness_pct"),
                    "message": f"PDF report generated. Download: {download_url}"
                }, indent=2)
            }]
        }

    except ImportError as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": f"ReportLab not installed. Install with: pip install reportlab. Details: {str(e)}"
                }, indent=2)
            }],
            "is_error": True
        }
    except Exception as e:
        import traceback
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="check_workflow_stage",
    description="Check if a workflow stage is complete by looking for stage completion file. Use this before starting a new stage to verify prerequisites are met.",
    input_schema={
        "session_context": str,
        "stage_name": str
    }
)
async def check_workflow_stage(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Check if a workflow stage is complete.
    
    Args:
        session_context: Path to session context directory
        stage_name: Name of stage to check (parse, batch, classify, aggregate, report)
    
    Returns:
        Stage completion status and metadata
    """
    try:
        stage_file = Path(args["session_context"]) / f"stage_{args['stage_name']}_complete.json"
        
        if stage_file.exists():
            with open(stage_file, 'r') as f:
                state = json.load(f)
            return {
                "content": [{
                    "type": "text",
                    "text": json.dumps({
                        "complete": True,
                        "stage": args["stage_name"],
                        "state": state
                    }, indent=2)
                }]
            }
        else:
            return {
                "content": [{
                    "type": "text",
                    "text": json.dumps({
                        "complete": False,
                        "stage": args["stage_name"],
                        "message": f"Stage {args['stage_name']} not yet complete"
                    }, indent=2)
                }]
            }
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "complete": False,
                    "error": str(e)
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="mark_stage_complete",
    description="Mark a workflow stage as complete by writing a completion file with metadata. Use this after successfully completing a stage.",
    input_schema={
        "session_context": str,
        "stage_name": str,
        "output_files": list,
        "metadata": dict
    }
)
async def mark_stage_complete(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Mark a workflow stage as complete.
    
    Args:
        session_context: Path to session context directory
        stage_name: Name of stage (parse, batch, classify, aggregate, report)
        output_files: List of files created by this stage
        metadata: Additional metadata about the stage completion
    
    Returns:
        Success confirmation
    """
    try:
        from datetime import datetime
        
        stage_file = Path(args["session_context"]) / f"stage_{args['stage_name']}_complete.json"
        
        state = {
            "stage": args["stage_name"],
            "status": "complete",
            "timestamp": datetime.now().isoformat(),
            "output_files": args.get("output_files", []),
            "metadata": args.get("metadata", {})
        }
        
        with open(stage_file, 'w') as f:
            json.dump(state, f, indent=2)
        
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": True,
                    "message": f"Stage {args['stage_name']} marked complete",
                    "state_file": str(stage_file)
                }, indent=2)
            }]
        }
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e)
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="wait_for_batch_file",
    description="Wait for a batch classification file to be created by a subagent. Polls for the file with timeout. Use this after spawning a batch-processor subagent.",
    input_schema={
        "session_context": str,
        "batch_number": int,
        "timeout_seconds": int
    }
)
async def wait_for_batch_file(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Wait for a batch classification file to be created.
    
    Args:
        session_context: Path to session context directory
        batch_number: Batch number to wait for
        timeout_seconds: Maximum time to wait (default: 120)
    
    Returns:
        File contents when ready, or timeout error
    """
    import asyncio
    
    try:
        batch_file = Path(args["session_context"]) / f"batch_{args['batch_number']}_elements.json"
        timeout = args.get("timeout_seconds", 120)
        poll_interval = 5
        elapsed = 0
        
        while elapsed < timeout:
            if batch_file.exists():
                # Wait a bit more to ensure file is fully written
                await asyncio.sleep(2)
                
                try:
                    with open(batch_file, 'r') as f:
                        data = json.load(f)
                    
                    # Validate it's a proper classification output
                    if isinstance(data, list) and len(data) > 0:
                        return {
                            "content": [{
                                "type": "text",
                                "text": json.dumps({
                                    "success": True,
                                    "batch_number": args["batch_number"],
                                    "element_count": len(data),
                                    "file_path": str(batch_file),
                                    "sample_element": data[0] if data else None
                                }, indent=2)
                            }]
                        }
                except json.JSONDecodeError:
                    pass  # File still being written
            
            await asyncio.sleep(poll_interval)
            elapsed += poll_interval
        
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": f"Timeout waiting for batch {args['batch_number']} after {timeout} seconds",
                    "expected_file": str(batch_file)
                }, indent=2)
            }],
            "is_error": True
        }
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e)
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="aggregate_batch_results",
    description="Combine all batch classification files into a single aggregated file. Validates that all batches are present and have consistent schema. Use this before generating reports.",
    input_schema={
        "session_context": str,
        "total_batches": int,
        "output_file": str
    }
)
async def aggregate_batch_results(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Aggregate all batch classification results into one file.
    
    Args:
        session_context: Path to session context directory
        total_batches: Expected number of batches
        output_file: Path to output aggregated file
    
    Returns:
        Aggregation results with element counts and validation status
    """
    try:
        import glob
        
        session_path = Path(args["session_context"])
        all_elements = []
        batch_stats = []
        missing_batches = []
        validation_errors = []
        
        for batch_num in range(1, args["total_batches"] + 1):
            batch_file = session_path / f"batch_{batch_num}_elements.json"
            
            if not batch_file.exists():
                missing_batches.append(batch_num)
                continue
            
            try:
                with open(batch_file, 'r') as f:
                    batch_data = json.load(f)
                
                if not isinstance(batch_data, list):
                    validation_errors.append(f"Batch {batch_num}: not a list")
                    continue
                
                # Validate and normalize field names
                for elem in batch_data:
                    # Normalize guid -> global_id
                    if 'guid' in elem and 'global_id' not in elem:
                        elem['global_id'] = elem.pop('guid')
                    
                    # Validate required fields
                    required = ['global_id', 'ifc_type']
                    for field in required:
                        if field not in elem:
                            validation_errors.append(f"Batch {batch_num}: missing {field}")
                
                all_elements.extend(batch_data)
                batch_stats.append({
                    "batch": batch_num,
                    "elements": len(batch_data)
                })
                
            except json.JSONDecodeError as e:
                validation_errors.append(f"Batch {batch_num}: invalid JSON - {e}")
        
        if missing_batches:
            return {
                "content": [{
                    "type": "text",
                    "text": json.dumps({
                        "success": False,
                        "error": f"Missing batches: {missing_batches}",
                        "found_batches": [s["batch"] for s in batch_stats]
                    }, indent=2)
                }],
                "is_error": True
            }
        
        # Write aggregated file
        output_path = Path(args["output_file"])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        aggregated = {
            "elements": all_elements,
            "total_count": len(all_elements),
            "batch_stats": batch_stats,
            "validation_warnings": validation_errors if validation_errors else None
        }
        
        with open(output_path, 'w') as f:
            json.dump(aggregated, f, indent=2)
        
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": True,
                    "total_elements": len(all_elements),
                    "batches_processed": len(batch_stats),
                    "batch_stats": batch_stats,
                    "output_file": str(output_path),
                    "validation_warnings": validation_errors if validation_errors else None
                }, indent=2)
            }]
        }
        
    except Exception as e:
        import traceback
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e),
                    "traceback": traceback.format_exc()
                }, indent=2)
            }],
            "is_error": True
        }


@tool(
    name="get_workflow_status",
    description="Get the overall status of the workflow by checking all stage completion files. Useful for resuming interrupted workflows.",
    input_schema={
        "session_context": str
    }
)
async def get_workflow_status(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Get comprehensive workflow status.
    
    Args:
        session_context: Path to session context directory
    
    Returns:
        Status of all workflow stages and available files
    """
    try:
        import glob
        
        session_path = Path(args["session_context"])
        
        if not session_path.exists():
            return {
                "content": [{
                    "type": "text",
                    "text": json.dumps({
                        "exists": False,
                        "message": f"Session context directory does not exist: {session_path}"
                    }, indent=2)
                }]
            }
        
        # Check stage completion files
        stages = ["parse", "batch", "classify", "aggregate", "report"]
        stage_status = {}
        
        for stage in stages:
            stage_file = session_path / f"stage_{stage}_complete.json"
            if stage_file.exists():
                with open(stage_file, 'r') as f:
                    stage_status[stage] = json.load(f)
            else:
                stage_status[stage] = {"status": "not_started"}
        
        # Check for data files
        files = {
            "parsed_data": (session_path / "parsed_data.json").exists(),
            "batches": (session_path / "batches.json").exists(),
            "aggregated": (session_path / "all_classified_elements.json").exists(),
            "co2_report": (session_path / "co2_report.json").exists()
        }
        
        # Count batch files
        batch_files = list(session_path.glob("batch_*_elements.json"))
        files["batch_files_count"] = len(batch_files)
        files["batch_files"] = [f.name for f in batch_files]
        
        # Check for output files
        output_files = list(session_path.glob("*.xlsx")) + list(session_path.glob("*.pdf"))
        files["output_files"] = [f.name for f in output_files]
        
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "session_context": str(session_path),
                    "stages": stage_status,
                    "files": files,
                    "recommendation": _get_workflow_recommendation(stage_status, files)
                }, indent=2)
            }]
        }
        
    except Exception as e:
        return {
            "content": [{
                "type": "text",
                "text": json.dumps({
                    "success": False,
                    "error": str(e)
                }, indent=2)
            }],
            "is_error": True
        }


# ==================== CSV AGENT TOOLS ====================

@tool(
    name="parse_csv",
    description="Parse a CSV file and return its contents. Supports CSV files from uploads or file paths. Returns data in a format ready for spreadsheet display. Use this as the first step when working with CSV data.",
    input_schema={
        "source": str,  # "file_path", "url", or "inline_data"
        "file_path": str,  # Path to CSV file (for source="file_path")
        "url": str,  # URL to fetch CSV from (for source="url")
        "inline_data": str,  # Raw CSV string (for source="inline_data")
        "delimiter": str,  # Optional: delimiter character (default: auto-detect)
        "has_header": bool,  # Optional: whether first row is header (default: True)
        "preview_rows": int  # Optional: limit rows for preview (default: all)
    }
)
async def parse_csv(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Parse CSV data from various sources.

    Args:
        source: Data source type - "file_path", "url", or "inline_data"
        file_path: Path to CSV file
        url: URL to fetch CSV from
        inline_data: Raw CSV string
        delimiter: CSV delimiter (auto-detected if not provided)
        has_header: Whether first row contains headers
        preview_rows: Limit number of rows returned

    Returns:
        Parsed CSV data with metadata
    """
    import csv
    import io

    try:
        source_type = args.get("source", "file_path")
        delimiter = args.get("delimiter", None)
        has_header = args.get("has_header", True)
        preview_rows = args.get("preview_rows", None)

        csv_content = ""
        source_info = ""

        # Get CSV content based on source
        if source_type == "file_path":
            file_path = args.get("file_path", "")
            if not file_path:
                return {
                    "content": [{"type": "text", "text": json.dumps({
                        "success": False,
                        "error": "file_path is required when source='file_path'"
                    }, indent=2)}],
                    "is_error": True
                }

            resolved_path = _resolve_path(file_path)
            if not resolved_path.exists():
                return {
                    "content": [{"type": "text", "text": json.dumps({
                        "success": False,
                        "error": f"File not found: {resolved_path}"
                    }, indent=2)}],
                    "is_error": True
                }

            with open(resolved_path, 'r', encoding='utf-8-sig') as f:
                csv_content = f.read()
            source_info = str(resolved_path)

        elif source_type == "url":
            url = args.get("url", "")
            if not url:
                return {
                    "content": [{"type": "text", "text": json.dumps({
                        "success": False,
                        "error": "url is required when source='url'"
                    }, indent=2)}],
                    "is_error": True
                }

            import httpx
            async with httpx.AsyncClient() as client:
                response = await client.get(url, timeout=30.0)
                response.raise_for_status()
                csv_content = response.text
            source_info = url

        elif source_type == "inline_data":
            csv_content = args.get("inline_data", "")
            if not csv_content:
                return {
                    "content": [{"type": "text", "text": json.dumps({
                        "success": False,
                        "error": "inline_data is required when source='inline_data'"
                    }, indent=2)}],
                    "is_error": True
                }
            source_info = "inline"
        else:
            return {
                "content": [{"type": "text", "text": json.dumps({
                    "success": False,
                    "error": f"Invalid source type: {source_type}. Use 'file_path', 'url', or 'inline_data'"
                }, indent=2)}],
                "is_error": True
            }

        # Auto-detect delimiter if not provided
        if not delimiter:
            sniffer = csv.Sniffer()
            try:
                sample = csv_content[:4096]
                dialect = sniffer.sniff(sample, delimiters=',;\t|')
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ','  # Default to comma

        # Parse CSV
        reader = csv.reader(io.StringIO(csv_content), delimiter=delimiter)
        rows = list(reader)

        if not rows:
            return {
                "content": [{"type": "text", "text": json.dumps({
                    "success": False,
                    "error": "CSV file is empty"
                }, indent=2)}],
                "is_error": True
            }

        # Extract headers
        headers = rows[0] if has_header else [f"Column_{i+1}" for i in range(len(rows[0]))]
        data_rows = rows[1:] if has_header else rows

        # Apply preview limit
        if preview_rows and preview_rows > 0:
            data_rows = data_rows[:preview_rows]

        # Build result
        total_rows = len(rows) - (1 if has_header else 0)

        return {
            "content": [{"type": "text", "text": json.dumps({
                "success": True,
                "source": source_info,
                "delimiter": delimiter,
                "total_rows": total_rows,
                "total_columns": len(headers),
                "headers": headers,
                "preview_rows": len(data_rows),
                "data": data_rows,
                "column_info": [{"index": i, "name": h, "sample_values": [row[i] if i < len(row) else "" for row in data_rows[:5]]} for i, h in enumerate(headers)]
            }, indent=2)}]
        }

    except Exception as e:
        import traceback
        return {
            "content": [{"type": "text", "text": json.dumps({
                "success": False,
                "error": str(e),
                "traceback": traceback.format_exc()
            }, indent=2)}],
            "is_error": True
        }


@tool(
    name="analyze_csv",
    description="Analyze CSV data and generate statistics. Provides column types, value distributions, missing values, and basic statistics. Use after parse_csv to understand the data structure.",
    input_schema={
        "file_path": str,  # Path to CSV file
        "delimiter": str  # Optional: delimiter character
    }
)
async def analyze_csv(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Analyze CSV data and generate statistics.

    Args:
        file_path: Path to CSV file
        delimiter: CSV delimiter (auto-detected if not provided)

    Returns:
        Analysis results with statistics per column
    """
    import csv
    import io
    from collections import Counter

    try:
        file_path = args.get("file_path", "")
        delimiter = args.get("delimiter", ",")

        resolved_path = _resolve_path(file_path)
        if not resolved_path.exists():
            return {
                "content": [{"type": "text", "text": json.dumps({
                    "success": False,
                    "error": f"File not found: {resolved_path}"
                }, indent=2)}],
                "is_error": True
            }

        with open(resolved_path, 'r', encoding='utf-8-sig') as f:
            csv_content = f.read()

        # Parse CSV
        reader = csv.reader(io.StringIO(csv_content), delimiter=delimiter)
        rows = list(reader)

        if len(rows) < 2:
            return {
                "content": [{"type": "text", "text": json.dumps({
                    "success": False,
                    "error": "CSV file has insufficient data for analysis"
                }, indent=2)}],
                "is_error": True
            }

        headers = rows[0]
        data_rows = rows[1:]

        # Analyze each column
        column_analysis = []
        for col_idx, header in enumerate(headers):
            values = [row[col_idx] if col_idx < len(row) else "" for row in data_rows]
            non_empty = [v for v in values if v.strip()]

            # Detect type
            numeric_count = 0
            float_count = 0
            for v in non_empty:
                try:
                    float(v.replace(",", ""))
                    numeric_count += 1
                    if "." in v:
                        float_count += 1
                except ValueError:
                    pass

            if numeric_count > len(non_empty) * 0.8:
                col_type = "float" if float_count > numeric_count * 0.3 else "integer"
            else:
                col_type = "string"

            # Calculate stats
            stats = {
                "name": header,
                "index": col_idx,
                "type": col_type,
                "total_count": len(values),
                "non_empty_count": len(non_empty),
                "empty_count": len(values) - len(non_empty),
                "unique_count": len(set(non_empty))
            }

            # Numeric stats
            if col_type in ["integer", "float"]:
                try:
                    numeric_values = [float(v.replace(",", "")) for v in non_empty if v.strip()]
                    if numeric_values:
                        stats["min"] = min(numeric_values)
                        stats["max"] = max(numeric_values)
                        stats["mean"] = sum(numeric_values) / len(numeric_values)
                        stats["sum"] = sum(numeric_values)
                except (ValueError, TypeError):
                    pass

            # Value distribution (top 5)
            value_counts = Counter(non_empty)
            stats["top_values"] = [{"value": v, "count": c} for v, c in value_counts.most_common(5)]

            column_analysis.append(stats)

        return {
            "content": [{"type": "text", "text": json.dumps({
                "success": True,
                "file": str(resolved_path),
                "total_rows": len(data_rows),
                "total_columns": len(headers),
                "columns": column_analysis
            }, indent=2)}]
        }

    except Exception as e:
        import traceback
        return {
            "content": [{"type": "text", "text": json.dumps({
                "success": False,
                "error": str(e),
                "traceback": traceback.format_exc()
            }, indent=2)}],
            "is_error": True
        }


@tool(
    name="csv_to_spreadsheet",
    description="Load CSV data directly into the Spreadsheet Builder. This is the PRIMARY tool for displaying CSV data in the UI. The user can then edit and save the data. Use this after parse_csv or directly with a file path.",
    input_schema={
        "file_path": str,  # Path to CSV file
        "name": str,  # Spreadsheet name (default: filename)
        "delimiter": str,  # Optional: delimiter character
        "max_rows": int  # Optional: limit rows loaded (default: 1000)
    }
)
async def csv_to_spreadsheet(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Load CSV into the Spreadsheet Builder UI component.

    Args:
        file_path: Path to CSV file
        name: Spreadsheet name
        delimiter: CSV delimiter
        max_rows: Maximum rows to load

    Returns:
        Spreadsheet data for UI display
    """
    import csv
    import io

    try:
        file_path = args.get("file_path", "")
        name = args.get("name", "")
        delimiter = args.get("delimiter", ",")
        max_rows = args.get("max_rows", 1000)

        resolved_path = _resolve_path(file_path)
        if not resolved_path.exists():
            return {
                "content": [{"type": "text", "text": json.dumps({
                    "success": False,
                    "error": f"File not found: {resolved_path}"
                }, indent=2)}],
                "is_error": True
            }

        # Default name from filename
        if not name:
            name = resolved_path.stem.replace("_", " ").title()

        with open(resolved_path, 'r', encoding='utf-8-sig') as f:
            csv_content = f.read()

        # Auto-detect delimiter if comma doesn't work well
        if delimiter == ",":
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(csv_content[:4096], delimiters=',;\t|')
                delimiter = dialect.delimiter
            except csv.Error:
                pass

        # Parse CSV
        reader = csv.reader(io.StringIO(csv_content), delimiter=delimiter)
        rows = list(reader)

        if not rows:
            return {
                "content": [{"type": "text", "text": json.dumps({
                    "success": False,
                    "error": "CSV file is empty"
                }, indent=2)}],
                "is_error": True
            }

        # Headers from first row
        headers = rows[0]
        data_rows = rows[1:max_rows + 1] if max_rows else rows[1:]

        # Include headers as first row for jspreadsheet (it expects data with header row)
        full_data = [headers] + data_rows

        # Build column definitions
        columns = []
        for i, header in enumerate(headers):
            # Estimate width based on header and content
            max_len = len(header)
            for row in data_rows[:20]:  # Sample first 20 rows
                if i < len(row):
                    max_len = max(max_len, len(str(row[i])))
            width = min(max(80, max_len * 10), 250)  # Between 80 and 250
            columns.append({"title": header, "width": width})

        total_rows = len(rows) - 1
        loaded_rows = len(data_rows)

        spreadsheet_data = {
            "type": "spreadsheet",
            "name": name,
            "data": full_data,
            "columns": columns,
            "row_count": loaded_rows + 1,  # +1 for header
            "column_count": len(columns)
        }

        return {
            "content": [{"type": "text", "text": json.dumps({
                "success": True,
                "spreadsheet": spreadsheet_data,
                "source_file": str(resolved_path),
                "total_rows_in_file": total_rows,
                "rows_loaded": loaded_rows,
                "truncated": total_rows > loaded_rows,
                "message": f"CSV loaded into Spreadsheet Builder: '{name}' with {loaded_rows} rows. Switch to the Spreadsheet view to see and edit the data."
            }, indent=2)}]
        }

    except Exception as e:
        import traceback
        return {
            "content": [{"type": "text", "text": json.dumps({
                "success": False,
                "error": str(e),
                "traceback": traceback.format_exc()
            }, indent=2)}],
            "is_error": True
        }


@tool(
    name="transform_csv",
    description="Transform CSV data: filter rows, select columns, sort, or aggregate. Returns transformed data ready for spreadsheet display.",
    input_schema={
        "file_path": str,  # Path to CSV file
        "columns": list,  # Optional: columns to include (by name or index)
        "filter_column": str,  # Optional: column to filter on
        "filter_value": str,  # Optional: value to filter for
        "filter_operator": str,  # Optional: "equals", "contains", "greater", "less"
        "sort_column": str,  # Optional: column to sort by
        "sort_ascending": bool,  # Optional: sort direction
        "group_by": str,  # Optional: column to group by
        "aggregate": str  # Optional: aggregation function ("sum", "count", "avg", "min", "max")
    }
)
async def transform_csv(args: Dict[str, Any]) -> Dict[str, Any]:
    """
    Transform CSV data with filtering, sorting, and aggregation.

    Args:
        file_path: Path to CSV file
        columns: Columns to include
        filter_column: Column to filter on
        filter_value: Value to filter for
        filter_operator: Filter comparison operator
        sort_column: Column to sort by
        sort_ascending: Sort direction
        group_by: Column to group by for aggregation
        aggregate: Aggregation function

    Returns:
        Transformed data ready for spreadsheet
    """
    import csv
    import io
    from collections import defaultdict

    try:
        file_path = args.get("file_path", "")
        select_columns = args.get("columns", None)
        filter_column = args.get("filter_column", None)
        filter_value = args.get("filter_value", None)
        filter_operator = args.get("filter_operator", "equals")
        sort_column = args.get("sort_column", None)
        sort_ascending = args.get("sort_ascending", True)
        group_by = args.get("group_by", None)
        aggregate = args.get("aggregate", None)

        resolved_path = _resolve_path(file_path)
        if not resolved_path.exists():
            return {
                "content": [{"type": "text", "text": json.dumps({
                    "success": False,
                    "error": f"File not found: {resolved_path}"
                }, indent=2)}],
                "is_error": True
            }

        with open(resolved_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            rows = list(reader)

        # Filter rows
        if filter_column and filter_value is not None and filter_column in headers:
            filtered_rows = []
            for row in rows:
                val = row.get(filter_column, "")
                try:
                    if filter_operator == "equals":
                        if str(val).lower() == str(filter_value).lower():
                            filtered_rows.append(row)
                    elif filter_operator == "contains":
                        if str(filter_value).lower() in str(val).lower():
                            filtered_rows.append(row)
                    elif filter_operator == "greater":
                        if float(val) > float(filter_value):
                            filtered_rows.append(row)
                    elif filter_operator == "less":
                        if float(val) < float(filter_value):
                            filtered_rows.append(row)
                except (ValueError, TypeError):
                    pass
            rows = filtered_rows

        # Group by and aggregate
        if group_by and group_by in headers and aggregate:
            grouped = defaultdict(list)
            for row in rows:
                key = row.get(group_by, "")
                grouped[key].append(row)

            result_rows = []
            numeric_cols = [h for h in headers if h != group_by]

            for key, group_rows in grouped.items():
                result_row = {group_by: key}
                for col in numeric_cols:
                    try:
                        values = [float(r.get(col, 0)) for r in group_rows]
                        if aggregate == "sum":
                            result_row[col] = sum(values)
                        elif aggregate == "count":
                            result_row[col] = len(values)
                        elif aggregate == "avg":
                            result_row[col] = sum(values) / len(values) if values else 0
                        elif aggregate == "min":
                            result_row[col] = min(values) if values else 0
                        elif aggregate == "max":
                            result_row[col] = max(values) if values else 0
                    except (ValueError, TypeError):
                        result_row[col] = group_rows[0].get(col, "")
                result_rows.append(result_row)
            rows = result_rows

        # Sort
        if sort_column and sort_column in headers:
            try:
                rows.sort(key=lambda r: float(r.get(sort_column, 0)), reverse=not sort_ascending)
            except (ValueError, TypeError):
                rows.sort(key=lambda r: str(r.get(sort_column, "")), reverse=not sort_ascending)

        # Select columns
        if select_columns:
            # Handle column names or indices
            selected_headers = []
            for col in select_columns:
                if isinstance(col, int) and col < len(headers):
                    selected_headers.append(headers[col])
                elif col in headers:
                    selected_headers.append(col)
            headers = selected_headers if selected_headers else headers

        # Convert to 2D array for spreadsheet
        data = [headers]
        for row in rows:
            data.append([str(row.get(h, "")) for h in headers])

        columns = [{"title": h, "width": 120} for h in headers]

        spreadsheet_data = {
            "type": "spreadsheet",
            "name": f"Transformed: {resolved_path.stem}",
            "data": data,
            "columns": columns,
            "row_count": len(data),
            "column_count": len(columns)
        }

        return {
            "content": [{"type": "text", "text": json.dumps({
                "success": True,
                "spreadsheet": spreadsheet_data,
                "original_rows": len(rows),
                "message": f"Transformed data ready. Switch to Spreadsheet view to see results."
            }, indent=2)}]
        }

    except Exception as e:
        import traceback
        return {
            "content": [{"type": "text", "text": json.dumps({
                "success": False,
                "error": str(e),
                "traceback": traceback.format_exc()
            }, indent=2)}],
            "is_error": True
        }


def _get_workflow_recommendation(stages: dict, files: dict) -> str:
    """Generate a recommendation for the next workflow step."""
    if stages.get("report", {}).get("status") == "complete":
        return "Workflow complete. Output files are ready for download."
    
    if stages.get("aggregate", {}).get("status") == "complete":
        return "Run report generation stage."
    
    if files.get("batch_files_count", 0) > 0:
        return f"Found {files['batch_files_count']} batch files. Run aggregate_batch_results to combine them."
    
    if stages.get("batch", {}).get("status") == "complete":
        return "Run classification stage with batch-processor subagents."
    
    if stages.get("parse", {}).get("status") == "complete":
        return "Run batch preparation stage."
    
    if files.get("parsed_data"):
        return "Parsed data exists. Mark parse stage complete and continue."
    
    return "Start with parsing the IFC file."


def create_ifc_tools_server():
    """
    Create MCP server with all IFC analysis and CSV tools.

    Tools become available as:
    - mcp__ifc__parse_ifc_file
    - mcp__ifc__prepare_batches
    - mcp__ifc__calculate_co2
    - mcp__ifc__generate_spreadsheet (PREFERRED for tabular data)
    - mcp__ifc__generate_excel_report (only for explicit download requests)
    - mcp__ifc__generate_pdf_report
    - mcp__ifc__generate_presentation
    - mcp__ifc__check_workflow_stage
    - mcp__ifc__mark_stage_complete
    - mcp__ifc__wait_for_batch_file
    - mcp__ifc__aggregate_batch_results
    - mcp__ifc__get_workflow_status
    - mcp__ifc__parse_csv (CSV Agent)
    - mcp__ifc__analyze_csv (CSV Agent)
    - mcp__ifc__csv_to_spreadsheet (CSV Agent - PRIMARY for loading CSV to UI)
    - mcp__ifc__transform_csv (CSV Agent)
    """
    return create_sdk_mcp_server(
        name="ifc",
        version="1.0.0",
        tools=[
            parse_ifc_file,
            prepare_batches,
            calculate_co2,
            generate_spreadsheet,  # PREFERRED for tabular data - must come before generate_excel_report
            generate_excel_report,
            generate_pdf_report,
            generate_presentation,
            check_workflow_stage,
            mark_stage_complete,
            wait_for_batch_file,
            aggregate_batch_results,
            get_workflow_status,
            # CSV Agent tools
            parse_csv,
            analyze_csv,
            csv_to_spreadsheet,
            transform_csv
        ]
    )

